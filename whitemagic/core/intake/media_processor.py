"""MediaProcessor Protocol — Pluggable Multimodal Intake (v14.2).
================================================================
Defines a Protocol for media-type-specific processors that extract
text, metadata, and holographic biases from non-text file types.

Each processor:
  1. Declares which file extensions it can handle
  2. Extracts text content (for embedding + FTS)
  3. Extracts structured metadata
  4. Provides holographic coordinate biases per media type

Processors are registered in a chain; the first matching processor wins.

Usage:
    from whitemagic.core.intake.media_processor import get_processor_chain
    chain = get_processor_chain()
    result = chain.process(Path("paper.pdf"))
    if result:
        print(result.text, result.metadata, result.holographic_bias)
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Protocol, runtime_checkable

logger = logging.getLogger(__name__)


@dataclass
class ProcessedMedia:
    """Result of processing a media file."""

    text: str                                  # Extracted text content
    metadata: dict[str, Any] = field(default_factory=dict)  # Structured metadata
    holographic_bias: dict[str, float] = field(default_factory=dict)  # XYZW biases
    media_type: str = "unknown"                # e.g., "pdf", "image", "audio"
    source_path: str = ""
    pages: int = 0                             # For documents
    duration_seconds: float = 0.0              # For audio/video
    dimensions: tuple[int, int] | None = None  # For images (width, height)


@runtime_checkable
class MediaProcessor(Protocol):
    """Protocol for media-type-specific content extraction."""

    def can_process(self, path: Path) -> bool:
        """Check if this processor can handle the given file."""
        ...

    def supported_extensions(self) -> set[str]:
        """Return the set of file extensions this processor handles."""
        ...

    def extract(self, path: Path) -> ProcessedMedia | None:
        """Extract text, metadata, and holographic biases from a file.

        Returns None if extraction fails.
        """
        ...


# ---------------------------------------------------------------------------
# PDF Processor
# ---------------------------------------------------------------------------

class PDFProcessor:
    """Extract text and metadata from PDF files using PyMuPDF (fitz)."""

    _available: bool | None = None

    def can_process(self, path: Path) -> bool:
        return path.suffix.lower() == ".pdf" and self._check_available()

    def supported_extensions(self) -> set[str]:
        return {".pdf"}

    def _check_available(self) -> bool:
        if self._available is not None:
            return self._available
        try:
            import fitz  # noqa: F401  (PyMuPDF)
            self._available = True
        except ImportError:
            self._available = False
            logger.debug("PyMuPDF (fitz) not installed — PDF processing unavailable")
        return self._available

    def extract(self, path: Path) -> ProcessedMedia | None:
        if not self.can_process(path):
            return None
        try:
            import fitz

            doc = fitz.open(str(path))
            pages_text: list[str] = []
            for page in doc:
                pages_text.append(page.get_text())  # type: ignore[union-attr]

            full_text = "\n\n".join(pages_text)
            metadata = doc.metadata or {}
            doc.close()

            # Truncate to 50K chars for memory storage
            if len(full_text) > 50000:
                full_text = full_text[:50000] + "\n\n[... truncated ...]"

            return ProcessedMedia(
                text=full_text,
                metadata={
                    "pdf_title": metadata.get("title", ""),
                    "pdf_author": metadata.get("author", ""),
                    "pdf_subject": metadata.get("subject", ""),
                    "pdf_pages": len(pages_text),
                    "pdf_creator": metadata.get("creator", ""),
                },
                holographic_bias={
                    "x": -0.3,   # Documents tend toward logical
                    "y": 0.2,    # Documents tend toward abstract/macro
                    "z": -0.2,   # Documents are often historical
                    "w": 0.3,    # Documents carry weight
                },
                media_type="pdf",
                source_path=str(path),
                pages=len(pages_text),
            )
        except Exception as e:
            logger.warning(f"PDF extraction failed for {path}: {e}")
            return None


# ---------------------------------------------------------------------------
# DOCX Processor
# ---------------------------------------------------------------------------

class DocxProcessor:
    """Extract text from DOCX files using python-docx."""

    _available: bool | None = None

    def can_process(self, path: Path) -> bool:
        return path.suffix.lower() == ".docx" and self._check_available()

    def supported_extensions(self) -> set[str]:
        return {".docx"}

    def _check_available(self) -> bool:
        if self._available is not None:
            return self._available
        try:
            import docx  # noqa: F401
            self._available = True
        except ImportError:
            self._available = False
            logger.debug("python-docx not installed — DOCX processing unavailable")
        return self._available

    def extract(self, path: Path) -> ProcessedMedia | None:
        if not self.can_process(path):
            return None
        try:
            import docx

            doc = docx.Document(str(path))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            full_text = "\n\n".join(paragraphs)

            if len(full_text) > 50000:
                full_text = full_text[:50000] + "\n\n[... truncated ...]"

            return ProcessedMedia(
                text=full_text,
                metadata={"docx_paragraphs": len(paragraphs)},
                holographic_bias={
                    "x": -0.2,
                    "y": 0.1,
                    "z": -0.1,
                    "w": 0.2,
                },
                media_type="docx",
                source_path=str(path),
            )
        except Exception as e:
            logger.warning(f"DOCX extraction failed for {path}: {e}")
            return None


# ---------------------------------------------------------------------------
# Image Processor
# ---------------------------------------------------------------------------

class ImageProcessor:
    """Extract text descriptions from images using BLIP-2 or Ollama vision."""

    _available: bool | None = None

    def can_process(self, path: Path) -> bool:
        return path.suffix.lower() in self.supported_extensions() and self._check_available()

    def supported_extensions(self) -> set[str]:
        return {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".tiff"}

    def _check_available(self) -> bool:
        if self._available is not None:
            return self._available
        # Check for Pillow (minimum requirement)
        try:
            from PIL import Image  # noqa: F401
            self._available = True
        except ImportError:
            self._available = False
            logger.debug("Pillow not installed — image processing unavailable")
        return self._available

    def extract(self, path: Path) -> ProcessedMedia | None:
        if not self.can_process(path):
            return None
        try:
            from PIL import Image

            img = Image.open(str(path))
            width, height = img.size
            mode = img.mode
            img_format = img.format or path.suffix.lstrip(".")

            # Try Ollama vision for captioning (local, no API key needed)
            caption = self._caption_ollama(path)

            # Fallback: basic metadata description
            if not caption:
                caption = self._caption_basic(path, width, height, mode, img_format)

            return ProcessedMedia(
                text=caption,
                metadata={
                    "image_width": width,
                    "image_height": height,
                    "image_mode": mode,
                    "image_format": img_format,
                    "captioning_method": "ollama" if "ollama" not in caption[:20].lower() else "basic",
                },
                holographic_bias={
                    "x": 0.3,    # Images are emotional/perceptual
                    "y": -0.3,   # Images are concrete/micro
                    "z": 0.0,    # Timeless by default
                    "w": 0.2,    # Moderate weight
                },
                media_type="image",
                source_path=str(path),
                dimensions=(width, height),
            )
        except Exception as e:
            logger.warning(f"Image extraction failed for {path}: {e}")
            return None

    def _caption_ollama(self, path: Path) -> str:
        """Use Ollama vision model for image captioning."""
        try:
            import base64
            import urllib.request

            from whitemagic.utils.fast_json import dumps_str as _json_dumps
            from whitemagic.utils.fast_json import loads as _json_loads

            # Read and base64-encode the image
            image_data = base64.b64encode(path.read_bytes()).decode("utf-8")

            payload = _json_dumps({
                "model": "llava",  # Common vision model
                "prompt": "Describe this image in detail. What do you see?",
                "images": [image_data],
                "stream": False,
            }).encode("utf-8")

            req = urllib.request.Request(
                "http://localhost:11434/api/generate",
                data=payload,
                headers={"Content-Type": "application/json"},
            )
            with urllib.request.urlopen(req, timeout=30) as resp:
                result = _json_loads(resp.read().decode("utf-8"))
                return str(result.get("response", ""))
        except Exception:
            return ""

    def _caption_basic(
        self, path: Path, width: int, height: int, mode: str, fmt: str,
    ) -> str:
        """Basic metadata-based description when no vision model available."""
        name = path.stem.replace("_", " ").replace("-", " ")
        return (
            f"Image: {name}\n"
            f"Format: {fmt}, {width}x{height} pixels, mode: {mode}\n"
            f"File: {path.name}"
        )


# ---------------------------------------------------------------------------
# Audio Processor
# ---------------------------------------------------------------------------

class AudioProcessor:
    """Transcribe audio files using Whisper."""

    _available: bool | None = None

    def can_process(self, path: Path) -> bool:
        return path.suffix.lower() in self.supported_extensions() and self._check_available()

    def supported_extensions(self) -> set[str]:
        return {".mp3", ".wav", ".flac", ".ogg", ".m4a", ".wma"}

    def _check_available(self) -> bool:
        if self._available is not None:
            return self._available
        try:
            import whisper  # noqa: F401
            self._available = True
        except ImportError:
            # Try faster-whisper as alternative
            try:
                from faster_whisper import WhisperModel  # noqa: F401
                self._available = True
            except ImportError:
                self._available = False
                logger.debug("Neither whisper nor faster-whisper installed — audio processing unavailable")
        return self._available

    def extract(self, path: Path) -> ProcessedMedia | None:
        if not self.can_process(path):
            return None
        try:
            text, duration = self._transcribe(path)
            if not text:
                return None

            if len(text) > 50000:
                text = text[:50000] + "\n\n[... truncated ...]"

            return ProcessedMedia(
                text=text,
                metadata={
                    "audio_duration_seconds": duration,
                    "audio_format": path.suffix.lstrip("."),
                },
                holographic_bias={
                    "x": 0.2,    # Audio tends emotional
                    "y": -0.1,   # Often concrete/specific
                    "z": 0.0,    # Varies
                    "w": 0.2,    # Moderate weight
                },
                media_type="audio",
                source_path=str(path),
                duration_seconds=duration,
            )
        except Exception as e:
            logger.warning(f"Audio extraction failed for {path}: {e}")
            return None

    def _transcribe(self, path: Path) -> tuple[str, float]:
        """Transcribe audio. Returns (text, duration_seconds)."""
        # Try faster-whisper first (more efficient)
        try:
            from faster_whisper import WhisperModel
            model = WhisperModel("base", device="cpu", compute_type="int8")
            segments, info = model.transcribe(str(path))
            text = " ".join(seg.text for seg in segments)
            return text.strip(), info.duration
        except Exception:
            pass

        # Fallback to openai-whisper
        try:
            import whisper
            model = whisper.load_model("base")
            result = model.transcribe(str(path))
            return result["text"].strip(), result.get("duration", 0.0)
        except Exception:
            pass

        return "", 0.0


# ---------------------------------------------------------------------------
# Spreadsheet Processor
# ---------------------------------------------------------------------------

class SpreadsheetProcessor:
    """Extract text from CSV/XLSX files."""

    _available: bool | None = None

    def can_process(self, path: Path) -> bool:
        ext = path.suffix.lower()
        if ext == ".csv":
            return True  # csv is stdlib
        return ext in {".xlsx", ".xls"} and self._check_available()

    def supported_extensions(self) -> set[str]:
        return {".csv", ".xlsx", ".xls"}

    def _check_available(self) -> bool:
        if self._available is not None:
            return self._available
        try:
            import openpyxl  # noqa: F401
            self._available = True
        except ImportError:
            self._available = False
            logger.debug("openpyxl not installed — XLSX processing unavailable")
        return self._available

    def extract(self, path: Path) -> ProcessedMedia | None:
        ext = path.suffix.lower()
        try:
            if ext == ".csv":
                return self._extract_csv(path)
            elif ext in {".xlsx", ".xls"}:
                return self._extract_xlsx(path)
        except Exception as e:
            logger.warning(f"Spreadsheet extraction failed for {path}: {e}")
        return None

    def _extract_csv(self, path: Path) -> ProcessedMedia | None:
        import csv
        rows: list[str] = []
        with open(path, newline="", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= 1000:
                    rows.append("[... truncated at 1000 rows ...]")
                    break
                rows.append(" | ".join(row))

        text = "\n".join(rows)
        return ProcessedMedia(
            text=text,
            metadata={"csv_rows": len(rows), "format": "csv"},
            holographic_bias={"x": -0.5, "y": -0.2, "z": 0.0, "w": 0.2},
            media_type="csv",
            source_path=str(path),
        )

    def _extract_xlsx(self, path: Path) -> ProcessedMedia | None:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheets_text: list[str] = []
        total_rows = 0

        for sheet_name in wb.sheetnames[:5]:  # Max 5 sheets
            ws = wb[sheet_name]
            sheet_rows: list[str] = [f"## Sheet: {sheet_name}"]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 500:
                    sheet_rows.append("[... truncated at 500 rows ...]")
                    break
                cells = [str(c) if c is not None else "" for c in row]
                sheet_rows.append(" | ".join(cells))
                total_rows += 1
            sheets_text.append("\n".join(sheet_rows))

        wb.close()
        text = "\n\n".join(sheets_text)

        if len(text) > 50000:
            text = text[:50000] + "\n\n[... truncated ...]"

        return ProcessedMedia(
            text=text,
            metadata={"xlsx_sheets": len(wb.sheetnames), "xlsx_rows": total_rows},
            holographic_bias={"x": -0.5, "y": -0.2, "z": 0.0, "w": 0.3},
            media_type="xlsx",
            source_path=str(path),
        )


# ---------------------------------------------------------------------------
# Processor Chain
# ---------------------------------------------------------------------------

class ProcessorChain:
    """Ordered chain of media processors. First matching processor wins."""

    def __init__(self) -> None:
        self._processors: list[MediaProcessor] = []

    def register(self, processor: MediaProcessor) -> None:
        """Register a processor in the chain."""
        self._processors.append(processor)

    def process(self, path: Path) -> ProcessedMedia | None:
        """Process a file using the first matching processor."""
        for proc in self._processors:
            if proc.can_process(path):
                result = proc.extract(path)
                if result:
                    logger.info(
                        f"Processed {path.name} via {proc.__class__.__name__} "
                        f"({len(result.text)} chars)",
                    )
                    return result
        return None

    def supported_extensions(self) -> set[str]:
        """All extensions supported by registered processors."""
        exts: set[str] = set()
        for proc in self._processors:
            exts.update(proc.supported_extensions())
        return exts

    def available_processors(self) -> list[dict[str, Any]]:
        """List registered processors and their availability."""
        info = []
        for proc in self._processors:
            name = proc.__class__.__name__
            exts = proc.supported_extensions()
            available = hasattr(proc, "_available") and proc._available is not False  # type: ignore[union-attr]
            info.append({
                "name": name,
                "extensions": sorted(exts),
                "available": available,
            })
        return info


# ---------------------------------------------------------------------------
# Singleton
# ---------------------------------------------------------------------------

_chain: ProcessorChain | None = None


def get_processor_chain() -> ProcessorChain:
    """Get the global processor chain with all built-in processors registered."""
    global _chain
    if _chain is None:
        _chain = ProcessorChain()
        _chain.register(PDFProcessor())
        _chain.register(DocxProcessor())
        _chain.register(ImageProcessor())
        _chain.register(AudioProcessor())
        _chain.register(SpreadsheetProcessor())
    return _chain
